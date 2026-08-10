from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Annotated

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import case, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.security import get_current_user
from app.models.billing import Invoice, PortalDocument, Vendor
from app.models.enums import IssuePriority, IssueStatus
from app.models.issue import Issue
from app.models.project import Project
from app.models.rate import RateItem
from app.models.user import User
from app.schemas import DashboardStats
from app.services.issue_service import remaining_days

router = APIRouter(prefix="/analytics", tags=["analytics"])

ISSUE_HEADERS = [
    "ID",
    "Project",
    "Type",
    "Category",
    "Priority",
    "Status",
    "Chainage",
    "Deadline",
    "Remaining Days",
    "Reporter",
    "Contractor",
    "Created",
]


@router.get("/dashboard", response_model=DashboardStats)
async def dashboard_stats(
    db: Annotated[AsyncSession, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
):
    total_projects = (await db.execute(select(func.count(Project.id)))).scalar_one()
    total_issues = (await db.execute(select(func.count(Issue.id)))).scalar_one()

    status_rows = (
        await db.execute(select(Issue.status, func.count(Issue.id)).group_by(Issue.status))
    ).all()
    by_status = {s.value if hasattr(s, "value") else str(s): 0 for s in IssueStatus}
    for status, count in status_rows:
        by_status[status.value] = count

    issues = (await db.execute(select(Issue))).scalars().all()
    delayed = sum(1 for i in issues if remaining_days(i.deadline_date) < 0 and i.status != IssueStatus.CLOSED)

    closed = [i for i in issues if i.status == IssueStatus.CLOSED and i.verified_at and i.created_at]
    avg_resolution = None
    if closed:
        days = [(i.verified_at.date() - i.created_at.date()).days for i in closed]
        avg_resolution = round(sum(days) / len(days), 2)

    on_time = sum(1 for i in closed if i.verified_at and i.verified_at.date() <= i.deadline_date)
    compliance = round((on_time / len(closed)) * 100, 1) if closed else None

    contractor_rows = (
        await db.execute(
            select(
                Issue.assigned_contractor_id,
                func.count(Issue.id),
                func.sum(case((Issue.status == IssueStatus.CLOSED, 1), else_=0)),
            ).group_by(Issue.assigned_contractor_id)
        )
    ).all()
    contractor_performance = [
        {"contractor_id": cid, "total": total, "closed": int(closed_n or 0)}
        for cid, total, closed_n in contractor_rows
    ]

    surveyor_rows = (
        await db.execute(
            select(Issue.reported_by_id, func.count(Issue.id)).group_by(Issue.reported_by_id)
        )
    ).all()
    surveyor_performance = [{"surveyor_id": sid, "reported": count} for sid, count in surveyor_rows]

    total_invoices = (await db.execute(select(func.count(Invoice.id)))).scalar_one()
    inv_rows = (await db.execute(select(Invoice.status, func.count(Invoice.id)).group_by(Invoice.status))).all()
    invoices_by_status = {
        (s.value if hasattr(s, "value") else str(s)): c for s, c in inv_rows
    }
    total_documents = (await db.execute(select(func.count(PortalDocument.id)))).scalar_one()
    total_vendors = (await db.execute(select(func.count(Vendor.id)))).scalar_one()
    boq_total = (await db.execute(select(func.coalesce(func.sum(RateItem.boq_amount), 0)))).scalar_one()
    exec_total = (await db.execute(select(func.coalesce(func.sum(RateItem.executed_amount), 0)))).scalar_one()

    return DashboardStats(
        total_projects=total_projects,
        total_issues=total_issues,
        by_status=by_status,
        delayed_issues=delayed,
        avg_resolution_days=avg_resolution,
        timeline_compliance_pct=compliance,
        contractor_performance=contractor_performance,
        surveyor_performance=surveyor_performance,
        total_invoices=total_invoices,
        invoices_by_status=invoices_by_status,
        total_documents=total_documents,
        total_vendors=total_vendors,
        total_boq_amount=float(boq_total or 0),
        total_executed_amount=float(exec_total or 0),
    )


@router.get("/export/excel")
async def export_excel(
    db: Annotated[AsyncSession, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
    project_id: int | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    package_name: str | None = None,
    report_title: str | None = None,
    prepared_by: str | None = None,
    remarks: str | None = None,
    period_type: str | None = None,
):
    today = date.today()
    period = (period_type or "custom").strip().lower()
    if period == "daily":
        date_from = today
        date_to = today
        if not report_title:
            report_title = f"Daily Issues Report — {today.isoformat()}"
    elif period == "weekly":
        # Monday–Sunday of current week (ISO)
        date_from = today - timedelta(days=today.weekday())
        date_to = date_from + timedelta(days=6)
        if not report_title:
            report_title = f"Weekly Issues Report — {date_from.isoformat()} to {date_to.isoformat()}"
    else:
        period = "custom"

    stmt = select(Issue).order_by(Issue.id)
    if project_id is not None:
        stmt = stmt.where(Issue.project_id == project_id)
    if date_from is not None:
        stmt = stmt.where(Issue.created_at >= datetime.combine(date_from, datetime.min.time()))
    if date_to is not None:
        stmt = stmt.where(Issue.created_at <= datetime.combine(date_to, datetime.max.time()))
    issues = (await db.execute(stmt)).scalars().all()

    project_name = ""
    if project_id is not None:
        proj = (await db.execute(select(Project).where(Project.id == project_id))).scalar_one_or_none()
        project_name = proj.name if proj else str(project_id)

    wb = Workbook()
    meta = wb.active
    meta.title = "Report Details"
    meta.append(["Field", "Value"])
    meta.append(["Report Type", period.title()])
    meta.append(["Report Title", report_title or "RoadService Issues Report"])
    meta.append(["Project ID", project_id or ""])
    meta.append(["Project Name", project_name])
    meta.append(["Package / Stretch", package_name or ""])
    meta.append(["Period From", date_from.isoformat() if date_from else ""])
    meta.append(["Period To", date_to.isoformat() if date_to else ""])
    meta.append(["Prepared By", prepared_by or ""])
    meta.append(["Remarks", remarks or ""])
    meta.append(["Generated On", today.isoformat()])
    meta.append(["Row Count", len(issues)])

    # Summary sheet for daily/weekly MIS
    summary = wb.create_sheet("Summary")
    summary.append(["Status", "Count"])
    status_counts: dict[str, int] = {}
    for i in issues:
        key = i.status.value if hasattr(i.status, "value") else str(i.status)
        status_counts[key] = status_counts.get(key, 0) + 1
    for k, v in sorted(status_counts.items()):
        summary.append([k, v])
    summary.append([])
    summary.append(["Total issues", len(issues)])
    delayed = sum(
        1
        for i in issues
        if i.deadline_date and remaining_days(i.deadline_date) < 0 and (i.status.value if hasattr(i.status, "value") else str(i.status)) != "closed"
    )
    summary.append(["Delayed (open)", delayed])

    ws = wb.create_sheet("Issues")
    ws.append(ISSUE_HEADERS)
    for i in issues:
        ws.append(
            [
                i.id,
                i.project_id,
                i.issue_type,
                i.work_category,
                i.priority.value if hasattr(i.priority, "value") else i.priority,
                i.status.value if hasattr(i.status, "value") else i.status,
                i.chainage,
                i.deadline_date.isoformat() if i.deadline_date else "",
                remaining_days(i.deadline_date) if i.deadline_date else "",
                i.reported_by_id,
                i.assigned_contractor_id,
                i.created_at.isoformat() if i.created_at else "",
            ]
        )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    suffix = period if period != "custom" else today.isoformat()
    fname = f"roadservice_{period}_report_{suffix}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.post("/import/excel")
async def import_excel(
    db: Annotated[AsyncSession, Depends(get_db)],
    user: Annotated[User, Depends(get_current_user)],
    file: UploadFile = File(...),
):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Upload an Excel .xlsx file")
    raw = await file.read()
    try:
        wb = load_workbook(BytesIO(raw), data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {exc}") from exc

    ws = wb["Issues"] if "Issues" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel sheet is empty")

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    idx = {h: i for i, h in enumerate(headers)}
    required = ["ID", "Status", "Priority"]
    missing = [h for h in required if h not in idx]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Excel must include columns: {', '.join(required)}. Missing: {', '.join(missing)}",
        )

    updated = 0
    skipped = 0
    errors: list[str] = []
    for row_no, row in enumerate(rows[1:], start=2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        try:
            issue_id = int(row[idx["ID"]])
        except (TypeError, ValueError):
            skipped += 1
            errors.append(f"Row {row_no}: invalid ID")
            continue
        issue = (await db.execute(select(Issue).where(Issue.id == issue_id))).scalar_one_or_none()
        if not issue:
            skipped += 1
            errors.append(f"Row {row_no}: issue #{issue_id} not found")
            continue

        status_raw = str(row[idx["Status"]] or "").strip().lower().replace(" ", "_")
        priority_raw = str(row[idx["Priority"]] or "").strip().lower()
        try:
            if status_raw:
                issue.status = IssueStatus(status_raw)
            if priority_raw:
                issue.priority = IssuePriority(priority_raw)
        except ValueError as exc:
            skipped += 1
            errors.append(f"Row {row_no}: {exc}")
            continue

        if "Type" in idx and row[idx["Type"]]:
            issue.issue_type = str(row[idx["Type"]]).strip()
        if "Category" in idx and row[idx["Category"]]:
            issue.work_category = str(row[idx["Category"]]).strip()
        if "Chainage" in idx and row[idx["Chainage"]] is not None:
            issue.chainage = str(row[idx["Chainage"]]).strip()
        updated += 1

    await db.commit()
    return {
        "ok": True,
        "updated": updated,
        "skipped": skipped,
        "errors": errors[:20],
        "imported_by": user.full_name,
        "filename": file.filename,
    }


@router.get("/export/pdf")
async def export_pdf(
    db: Annotated[AsyncSession, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
):
    issues = (await db.execute(select(Issue).order_by(Issue.id).limit(200))).scalars().all()
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    styles = getSampleStyleSheet()
    story = [
        Paragraph("Road Issue Management — Issues Report", styles["Title"]),
        Spacer(1, 12),
        Paragraph(f"Generated: {date.today().isoformat()}", styles["Normal"]),
        Spacer(1, 18),
    ]
    data = [["ID", "Type", "Status", "Priority", "Deadline"]]
    for i in issues:
        data.append([str(i.id), i.issue_type, i.status.value, i.priority.value, i.deadline_date.isoformat()])
    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f7fa")]),
        ])
    )
    story.append(table)
    doc.build(story)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="issues_{date.today()}.pdf"'},
    )
